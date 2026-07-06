"""Document chunker — section-aware splitting for markdown, PDF, and Excel files."""

import re
from dataclasses import dataclass, field
from pathlib import Path


@dataclass
class Chunk:
    text: str
    metadata: dict = field(default_factory=dict)


def chunk_markdown(filepath: Path, max_tokens: int = 400, overlap_tokens: int = 50) -> list[Chunk]:
    """Split a markdown file on headers, keeping sections as atomic chunks.

    If a section exceeds max_tokens, it's split at paragraph boundaries.
    """
    text = filepath.read_text(encoding="utf-8")
    sections = _split_on_headers(text)

    doc_type = _infer_doc_type(filepath)
    station_id = _extract_station_id(text)

    chunks = []
    for section_header, section_body in sections:
        full_text = f"{section_header}\n\n{section_body}".strip()
        if not full_text:
            continue

        word_count = len(full_text.split())
        if word_count <= max_tokens:
            chunks.append(Chunk(
                text=full_text,
                metadata={
                    "source": filepath.name,
                    "doc_type": doc_type,
                    "section": section_header.strip("# ").strip(),
                    "station_id": station_id,
                },
            ))
        else:
            sub_chunks = _split_on_paragraphs(full_text, max_tokens, overlap_tokens)
            for i, sub in enumerate(sub_chunks):
                chunks.append(Chunk(
                    text=sub,
                    metadata={
                        "source": filepath.name,
                        "doc_type": doc_type,
                        "section": f"{section_header.strip('# ').strip()} (part {i+1})",
                        "station_id": station_id,
                    },
                ))

    return chunks


def chunk_pdf(filepath: Path, max_tokens: int = 400) -> list[Chunk]:
    """Extract text from PDF and chunk by pages/sections."""
    try:
        import pymupdf
    except ImportError:
        return [Chunk(text=f"[PDF not parseable — pymupdf not installed: {filepath.name}]",
                      metadata={"source": filepath.name, "doc_type": "drawing"})]

    doc = pymupdf.open(str(filepath))
    chunks = []
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text().strip()
        if not text:
            continue

        chunks.append(Chunk(
            text=text[:max_tokens * 5],  # rough limit
            metadata={
                "source": filepath.name,
                "doc_type": "drawing",
                "page": page_num + 1,
            },
        ))
    doc.close()
    return chunks


def chunk_excel(filepath: Path) -> list[Chunk]:
    """Extract data from Excel BOM files as structured text chunks."""
    try:
        import openpyxl
    except ImportError:
        return [Chunk(text=f"[Excel not parseable — openpyxl not installed: {filepath.name}]",
                      metadata={"source": filepath.name, "doc_type": "bom"})]

    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    chunks = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]

        # Chunk every 20 rows
        batch_size = 20
        for i in range(1, len(rows), batch_size):
            batch = rows[i:i + batch_size]
            lines = [" | ".join(headers)]
            lines.append("-" * 40)
            for row in batch:
                line = " | ".join(str(v) if v is not None else "" for v in row)
                lines.append(line)

            chunks.append(Chunk(
                text="\n".join(lines),
                metadata={
                    "source": filepath.name,
                    "doc_type": "bom",
                    "sheet": sheet_name,
                    "rows": f"{i}-{i + len(batch) - 1}",
                },
            ))

    wb.close()
    return chunks


def chunk_file(filepath: Path) -> list[Chunk]:
    """Auto-detect file type and chunk accordingly."""
    suffix = filepath.suffix.lower()
    if suffix == ".md":
        return chunk_markdown(filepath)
    elif suffix == ".pdf":
        return chunk_pdf(filepath)
    elif suffix in (".xlsx", ".xls"):
        return chunk_excel(filepath)
    else:
        text = filepath.read_text(encoding="utf-8", errors="ignore")
        return [Chunk(text=text, metadata={"source": filepath.name, "doc_type": "unknown"})]


def chunk_all_documents(docs_dir: Path) -> list[Chunk]:
    """Recursively chunk all documents in the RAG corpus directory."""
    all_chunks = []
    for filepath in sorted(docs_dir.rglob("*")):
        if filepath.is_file() and filepath.suffix.lower() in (".md", ".pdf", ".xlsx", ".xls"):
            chunks = chunk_file(filepath)
            all_chunks.extend(chunks)
    return all_chunks


def _split_on_headers(text: str) -> list[tuple[str, str]]:
    """Split markdown text on ## and # headers."""
    pattern = r"^(#{1,3}\s+.+)$"
    parts = re.split(pattern, text, flags=re.MULTILINE)

    sections = []
    current_header = ""
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if re.match(r"^#{1,3}\s+", part):
            current_header = part
        else:
            sections.append((current_header, part))

    return sections


def _split_on_paragraphs(text: str, max_tokens: int, overlap_tokens: int) -> list[str]:
    """Split text at paragraph boundaries to fit within token limit."""
    paragraphs = text.split("\n\n")
    chunks = []
    current = ""

    for para in paragraphs:
        candidate = f"{current}\n\n{para}".strip() if current else para
        if len(candidate.split()) <= max_tokens:
            current = candidate
        else:
            if current:
                chunks.append(current)
            current = para

    if current:
        chunks.append(current)

    return chunks


def _infer_doc_type(filepath: Path) -> str:
    """Infer document type from directory structure."""
    parts = filepath.parts
    for part in parts:
        if part == "sops":
            return "sop"
        elif part == "manuals":
            return "manual"
        elif part in ("specs", "process_specs", "product_specs"):
            return "spec"
        elif part == "troubleshooting_db":
            return "troubleshooting"
        elif part in ("fmea", "dfmea", "pfmea"):
            return "fmea"
        elif part == "case_studies":
            return "case_study"
        elif part == "drawings":
            return "drawing"
        elif part == "bom":
            return "bom"
        elif part == "pdca_records":
            return "pdca"
    return "unknown"


def _extract_station_id(text: str) -> str | None:
    """Try to extract station ID from document text."""
    match = re.search(r"(STP|WLD|PNT|ASM|QAT)-\d{2}-[A-Z]{3}", text)
    return match.group(0) if match else None
